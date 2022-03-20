import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill
from openpyxl.styles import colors
import re

def create_master_list(my_dir):
    p = Path(my_dir)
    return os.listdir(p)

def extract_date(list):
    date_list = []
    len_list = len(list)
    for i in range(0, len_list):
        date_list = date_list + [list[i][0:10]]
    return date_list

def extract_name(list):
    name_list = []
    len_list = len(list)
    for i in range(0, len_list):
        max_word = len(list[i])
        if max_word < 11:
            name_list = name_list + [list[i][10:max_word]] #only to be used if there are short file names - and if so, change 'if' on the next line to 'elif'
        elif list[i][10] == " ":
            name_list = name_list + [list[i][11:max_word]]
        else:
            name_list = name_list + [list[i][10:max_word]]
    return name_list


# the sorted_alphanumeric function is taken from https://stackoverflow.com/questions/4836710/is-there-a-built-in-function-for-string-natural-sort
def sorted_alphanumeric(data):
    convert = lambda text: int(text) if text.isdigit() else text.lower()
    alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key)]
    return sorted(data, key=alphanum_key)

def multiple_folders(input_dir, output_name):
    my_grey = openpyxl.styles.colors.Color(rgb='00D0CECE')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_grey)
    y = []
    serial_number = 1
    initial_value = 2
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet.cell(row=1, column=1).value = 'S/No.'
    sheet.cell(row=1, column=2).value = 'Date of Document (YYYY.MM.DD)'
    sheet.cell(row=1, column=3).value = 'Description of Document'
    for z in range(1,4):
        sheet.cell(row=1, column=z).font = Font(bold=True)
    sheet.cell(row=1, column=4).value = 'Old File name'
    sheet.cell(row=1, column=5).value = 'New File name'
    for folderName, subfolders, filenames in os.walk(input_dir):
        y = y + [folderName]
    length_y = len(y)
    y = y[1:length_y]
    y = sorted_alphanumeric(y)
    length_y_new = len(y)
    for i in range(0, length_y_new):
        p11 = Path(y[i])
        dir_path_extension = os.listdir(p11)
        dir_path = [os.path.splitext(x)[0] for x in dir_path_extension]
        my_date_list = extract_date(dir_path)
        my_name_list = extract_name(dir_path)
        length_list = len(my_date_list)
        sheet.cell(row=initial_value, column=1).value = os.path.basename(p11)
        sheet.cell(row=initial_value, column=1).alignment = Alignment(horizontal='center')
        sheet.cell(row=initial_value, column=1).font = Font(bold=True, italic=True)
        sheet.cell(row=initial_value, column=1).fill = my_fill
        sheet.merge_cells(start_row=initial_value, start_column=1, end_row=initial_value, end_column=3)
        for k in range(1,length_list+2):
            if k < length_list + 1:
                sheet.cell(row=k + initial_value, column=1).value = str(serial_number) + '.'
                sheet.cell(row=k + initial_value, column=2).value = my_date_list[k - 1]
                sheet.cell(row=k + initial_value, column=3).value = my_name_list[k - 1]
                sheet.cell(row=k + initial_value, column=4).value = dir_path_extension[k - 1]
                sheet.cell(row=k + initial_value, column=5).value = str(serial_number) + ". " + dir_path_extension[k - 1]
                serial_number = serial_number + 1
            else:
                initial_value = initial_value + k
    for y in range(1,initial_value):
        for u in range(1,4):
            sheet.cell(row=y, column=u).border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    wb.save(output_name)
